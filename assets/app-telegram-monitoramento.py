import asyncio
import logging
import os
from datetime import date, datetime
from datetime import datetime

from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, Updater, MessageHandler, filters, ConversationHandler
from telegram import InputFile

from driver import Driver
from send_email import send
from config.webdriver import iniciar_navegador
from config.gerar_pdf import gerar_pdf

import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ConversationHandler, MessageHandler, filters, CallbackContext
from warnings import filterwarnings
from telegram.warnings import PTBUserWarning
import asyncio
filterwarnings(action="ignore", message=r".*CallbackQueryHandler", category=PTBUserWarning)
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ConversationHandler, MessageHandler, filters, CallbackContext
from driver import Driver
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
import pandas as pd

DATEFORMAT = '%d-%m-%Y %H:%M:%S'
logging.basicConfig(filename='logfile.txt', level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s: %(message)s',
                    datefmt=DATEFORMAT)


class kintBot:
    # Token do KinBot Engenharia.
    # TOKEN = '6113497316:AAEivqHzmiiTMMZKI6RKCzQMhCoz62I4OEY'

    # Token do seu robô de testes.

    # TOKEN = '5960331687:AAGKEq6Psw3XsBa6nswouVZcQ1_I-mzssRM'
    # TOKEN = '5840284976:AAE0pGUW3v79DZtZ6cKiGlum83k4b5VwqPs'

    # TOKEN = '6302929514:AAGLTIMfWBaEzfGFRMex2hhvFgl6MXw_4BQ' #bfautorobot

    """MONITORAMENTO"""
    TOKEN = '6419590681:AAEbu154z9-3PaFlMaQF2q9HM3nyGotyEh8' # @ENGENHARIA_KINBOT
    
    """RELATORIOS DE GERAÇÃO MENSAL"""
    # TOKEN = '6643696736:AAGgLIixBc5j_9r6wv2EdgmaG8Na07pnnO4' # @ENGENHARIA_RELATORIOS_KINBOT


    #### PARA TESTES ######
    # TOKEN = '6530582523:AAF-nZl5q9Ws8JKrqLaGpngGfQ1sfF-8KyA' # @engenhariakinsolbot

    user_settings = {
        # 'username_solarz': 'engenharia@kinsolenergia.com.br',
        # 'userpasswd_solarz': '@Kinsol21',
        'username_solarz': 'juliag.kinsol@gmail.com',
        'userpasswd_solarz': 'Kinsol21',
        'username_crm': 'kinsol.servidor@gmail.com',
        'userpasswd_crm': 'kinsolbot10'
    }

    OPTION_EXTRACT, SELECT_PAGE, TEXTO_UM, TEXTO_DOIS, SELECT_EXECUTIVO = range(5)

    def datehours(self):
        return datetime.now().strftime(DATEFORMAT)

    async def send_long_messages(self, update, message, clients) -> None:
        client_message = message
        for client in clients:
            if len(client_message) > 4000:
                await update.message.reply_text(client_message)
                client_message = ''
                continue
            client_message += f'{client["name"]}\n'
        await update.message.reply_text(client_message)

                    
    async def start(self, update: Update, context) -> None:
            """Send a message when the command /start is issued."""
            user = update.message.from_user
            logging.info(f'Comando /start executado por {user.first_name} {user.last_name} - ID:{user.id}')
            # INFO
            await update.message.reply_text(f'Olá {user.first_name} {user.last_name} 👋, bom te ver!')
            # try:
            #     await self.enviar_relatorio_excel_monitoramento_geral(update)
            # except Exception as e: print(f'Erro: {e}')
    ###     monitoramento   #####################

    async def erro(erro):
            hora_atual = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log = f'>>> [{hora_atual}] [ERROR] {erro}'
            print(log)
    
    async def button_callback_monitoramento(self, update: Update, context: CallbackContext):
        
        query = update.callback_query
        user_choice = query.data

        if user_choice == 'engenharia':
            print(user_choice)
            usuario = 'engenharia'
            senha = 'senha'
            user = 'engenharia'
            #monitoramento(user, senha)
            await query.message.reply_text('🚀 Iniciando o monitoramento no perfil ENGENHARIA. \n\n🕐 Aguarde a extração ser concluída...\n🔒 Realizando Login no SolarZ ... ')

        elif user_choice == 'cancelar':
                
            await query.message.reply_text('✖️ Operação Cancelada! ✖️')
            return ConversationHandler.END
        
        if user_choice != 'cancelar':
                try:
                    await query.message.reply_text('🚀  Iniciando monitoramento.. Aguarde..')

                    """Reliza a captura de informações sobre desemepnho do cliente"""
                    driver = Driver(**self.user_settings)

                    try:
                        try:
                            response = await driver.monitoramento(user_choice, update)
                            await driver.driver_finish()

                        except Exception as e: print(e)
                        if response == True:
                            
                            await query.message.reply_text('✅  Monitoramento finalizado.\n\n📑 Planilha gerada com sucesso.')
                        try:
                                await self.enviar_relatorio_excel_monitoramento(update)
                                await self.enviar_relatorio_excel_monitoramento_geral(update)
                        except Exception as e: print(e)
                    except Exception as e: print(e)
                except Exception as e: print(e)
        
        return ConversationHandler.END
    
    async def get_option_executivos_monitoramento(self, update: Update, context: CallbackContext):
        keyboard = [
            [InlineKeyboardButton("ENGENHARIA", callback_data='engenharia')],
            # [InlineKeyboardButton("João Victor Nunes", callback_data='user1')],
            # [InlineKeyboardButton("Victor Leoni", callback_data='user2')],
            # [InlineKeyboardButton("Álvaro Nagay", callback_data='user3')],
            # [InlineKeyboardButton("TESTE", callback_data='user4')],
            
            [InlineKeyboardButton("CANCELAR", callback_data='cancelar')],

        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await update.message.reply_text("🚀  Vamos inicar o monitoramento..\n\n📑  Escolha com qual perfil deseja realizar o monitoramento... ", reply_markup=reply_markup)

        return self.SELECT_EXECUTIVO   
    
    async def enviar_relatorio_excel_monitoramento(self, update:Update) -> None:

        query = update.callback_query

        try:
            # Diretório atual do script
            relatorios_directory = os.path.dirname(os.path.abspath(__file__))

            # Caminho completo para o arquivo XLSX
            xlsx_file_path = os.path.join(relatorios_directory, 'monitoramento.xlsx')
            print(xlsx_file_path)
            try:
                # Verifica se o arquivo existe
                if os.path.exists(xlsx_file_path):
                    # Cria um objeto InputFile com o caminho do arquivo
                
                    with open(xlsx_file_path, 'rb') as arquivo:
                        try:
                            xlsx_input_file = InputFile(arquivo)
                            await query.message.reply_document(xlsx_input_file, caption='✔️📊 Planilha enviada com sucesso!')
                        except Exception as e: print('🚨    Erro ao enviar arquivo excel: ', e)

                    # await update.message.reply_document(xlsx_input_file, caption='✔️📊 Planilha enviada com sucesso!')
                else:
                     # Diretório atual do script
                    relatorios_directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

                    # Caminho completo para o arquivo XLSX
                    xlsx_file_path = os.path.join(relatorios_directory,'monitoramento.xlsx')
                    print(xlsx_file_path)
                     # Verifica se o arquivo existe
                    if os.path.exists(xlsx_file_path):
                        # Cria um objeto InputFile com o caminho do arquivo
                    
                        with open(xlsx_file_path, 'rb') as arquivo:
                            try:
                                xlsx_input_file = InputFile(arquivo)
                                await query.message.reply_document(xlsx_input_file, caption='✔️📊 Planilha enviada com sucesso!')
                            except Exception as e: print('🚨    Erro ao enviar arquivo excel: ', e)
                    else:
                            await query.message.reply_text("💀 Falha ao encontrar arquivo.")

            except Exception as e: print(e)
        except Exception as e:await query.message.reply_text((f'>>>> ERROR [Enviar_relatorio_monitoramento] \n>>{e} '))
    
    async def enviar_relatorio_excel_monitoramento_geral(self, update: Update) -> None:

        # Função para encontrar a coluna correspondente à data
        def encontrar_coluna(ws, data):
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == data:
                    return col
            return None

        # Caminho para os arquivos
        path = os.path.dirname(os.path.abspath(__file__))

        caminho_monitoramento = os.path.join(path, 'monitoramento.xlsx')
        caminho_monitoramento_geral = os.path.join(path, 'MONITORAMENTO_GERAL.xlsx')

        # Carregue o arquivo Excel em um DataFrame do pandas
        df = pd.read_excel(caminho_monitoramento)
        data_cadastrada = df.columns[2]  # A terceira coluna tem o índice 2

        # Crie uma lista de listas para armazenar os dados das usinas
        usinas = []

        # Itere pelas linhas do DataFrame e extraia os valores de Usina e Desempenho
        for index, row in df.iterrows():
            nome_usina = row['USINA']
            desempenho = row.iloc[2]  # Lê o valor da terceira coluna (índice 2)
            usina_info = [nome_usina, desempenho, data_cadastrada]
            usinas.append(usina_info)

        # Carregue o arquivo Excel em um DataFrame do pandas usando openpyxl
        wb = load_workbook(caminho_monitoramento_geral)
        ws = wb.active

        # Obtenha a última data cadastrada (título da última coluna)
        ultima_data_cadastrada = ws.cell(row=1, column=ws.max_column).value

        # Verifique se a última data cadastrada é igual à data que você extraiu anteriormente
        if ultima_data_cadastrada != data_cadastrada:
            # Adicione uma nova coluna
            nova_coluna = ws.max_column + 1
            ws.cell(row=1, column=nova_coluna, value=data_cadastrada)
            print(f'Nova coluna adicionada para data: {data_cadastrada}')
            coluna = nova_coluna

        else:
             print(f'Coluna localizada para a data {data_cadastrada}  coluna={ws.max_column}')
             coluna = ws.max_column
        # print(f'usinas:\n\n{usinas}')
        # Itere pelas usinas e verifique se a usina já existe no arquivo
        for usina in usinas:
            nome_usina, desempenho, data_cadastrada = usina
            encontrada = False  # Variável para rastrear se a usina foi encontrada

            # Verifique se o nome da usina está presente na planilha
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                # print(f'Valor da row = {row} column= 1: {cell_value}')
                if cell_value is not None and cell_value == nome_usina:
                    # Usina encontrada, imprima a linha em que ela está
                    print(f'Usina "{nome_usina}" encontrada na linha {row}')
                    linha = row
                    encontrada = True
                    break
            if not encontrada:
                # A usina não está na planilha, então crie uma nova linha
                nova_linha = [nome_usina] + [None] * (ws.max_column - 1)
                ws.append(nova_linha)
                row = ws.max_row
                linha = row
                
                print(f'Nova linha criada para usina: "{nome_usina}" na linha {row}')
            # Agora que você tem a coluna e a linha corretas, insira o desempenho na célula apropriada
            if coluna and linha:
                ws.cell(row=linha, column=coluna, value=desempenho)
                print(f'Desempenho inserido para usina "{nome_usina}" na linha {linha}, coluna {coluna}')

        # Salve as alterações no arquivo
        wb.save(caminho_monitoramento_geral)
    
        query = update.callback_query

        print(caminho_monitoramento_geral)
        
        # Verifica se o arquivo existe
        if os.path.exists(caminho_monitoramento_geral):
                # Cria um objeto InputFile com o caminho do arquivo
            
                with open(caminho_monitoramento_geral, 'rb') as arquivo:
                    try:
                        xlsx_input_file = InputFile(arquivo)
                        await query.message.reply_document(caminho_monitoramento_geral, caption='✔️📊 Planilha GERAL enviada com sucesso!')
                    except Exception as e: print('🚨    Erro ao enviar arquivo excel: ', e)

                # await update.message.reply_document(xlsx_input_file, caption='✔️📊 Planilha enviada com sucesso!')

            
   
   
    #####################################

    def start_kin(self):
        print('>>> Lembre-se, todas as planilhas que serão manipuladas devem estar fechadas, caso contrário ocasionará em erros ao salvar o arquivo Excel.\n\n')
        self.application = Application.builder().token(self.TOKEN).build()

        self.application.add_handler(CommandHandler('start', self.start))
        self.application.add_handler(CommandHandler('e', self.enviar_relatorio_excel_monitoramento_geral))

        monitoramento_conv_handler = ConversationHandler(
            entry_points=[CommandHandler('monitoramento', self.get_option_executivos_monitoramento)],
            states={
                self.SELECT_EXECUTIVO: [CallbackQueryHandler(self.button_callback_monitoramento)],
            },
            fallbacks=[],
        )

        self.application.add_handler(monitoramento_conv_handler) ## /monitoramento

        self.application.run_polling()





if __name__ == '__main__':
    kinbot = kintBot()
    try:
        kinbot.start_kin()
    except Exception as e:
        print('>>> ', e)