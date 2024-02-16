import os
import asyncio
from datetime import date, datetime, timedelta
import datetime
## manipulação de excel
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from config.webdriver import iniciar_navegador
## webdrivers imports
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
# import autoit
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from unidecode import unidecode
import re
## SUPORTE CHROMIUM
from selenium.webdriver.chrome.service import Service as ChromiumService
from webdriver_manager.chrome import ChromeDriverManager
import psutil


## imports do monitoramento ###
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import IconSet, FormatObject, Rule
from webdriver_manager.chrome import ChromeDriverManager
import threading
import pyautogui
pyautogui.FAILSAFE = False
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class Driver:
    URLS_SITE = {
        'login_solarz': 'https://app.solarz.com.br/login?logout',
        'login_crm': 'https://app.kinsol.com.br',
        'users': 'https://app.kinsol.com.br/admin/users',
        'proposals': 'https://app.kinsol.com.br/admin/board',
    }

    file_name = ''
    initial_date = ''
    final_date = ''
    LIST_MONTHS = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril',
        'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro',
        'Outubro', 'Novembro', 'Dezembro'
    ]



    def __init__(self, username_solarz, userpasswd_solarz, username_crm, userpasswd_crm):
        try:
            self.driver = iniciar_navegador('Chrome')
            """Definir tamanho do navegador para todos
            os elementos serem renderizados na tela"""
            # self.driver.set_window_size(1820, 920)
            self.driver.maximize_window()

            self.username_solarz = username_solarz
            print(self.username_solarz)
            self.userpasswd_solarz = userpasswd_solarz
            self.username_crm = username_crm
            self.userpasswd_crm = userpasswd_crm
        except Exception as erro: print(erro)

    async def login_solarz(self) -> None:
        """Realiza login no site"""
        try:
            self.driver.get(self.URLS_SITE['login_solarz'])
            await asyncio.sleep(2)
        except:
            # quando receber o erro Timed Out, ele não consegue acessar o site pois o selenium para de responder
            # testar solução de fechar e iniciar nova instancia do driver 
            try:
                await self.driver_finish()
            except:
                print('>> Não conseguiu fechar, forçar kill process 💀')
                # Obter o PID do processo do navegador controlado pelo Selenium
  
            self.driver = iniciar_navegador('Chrome')
            self.driver.maximize_window()
            self.driver.get(self.URLS_SITE['login_solarz'])
            await asyncio.sleep(2)

        try:
            self.driver.find_element(
                By.ID, 'username').send_keys(self.username_solarz)
            self.driver.find_element(By.ID, 'password').send_keys(
                self.userpasswd_solarz)
            self.driver.find_element(
                By.CSS_SELECTOR, 'input[value="Entrar"]').click()
            await asyncio.sleep(5)
        except Exception as er:
            print('>>> [ERROR]: Erro ao localizar elemento')
            print(f'>>> [ERROR]83: {er}')

    async def login_crm(self) -> None:
        """Realiza login no site"""
        self.driver.get(self.URLS_SITE['login_crm'])
        await asyncio.sleep(2)

        try:
            self.driver.find_element(
                By.NAME, 'email').send_keys(self.username_crm)
            self.driver.find_element(
                By.NAME, 'password').send_keys(self.userpasswd_crm)
            self.driver.find_element(
                By.CSS_SELECTOR, 'button[type="submit"]').click()
        except NoSuchElementException as er:
            print('>>> [ERROR]: Erro ao localizar elemento')
            print(f'>>> [ERROR]99: {er}')

    async def set_period(self) -> None:
        # backup
        # try:
        #     current_month =  datetime.datetime.today().month
        #     # Obter a data do último dia do mês passado
        #     current_date =  datetime.datetime.today()
        #     first_day_current_month = date(current_date.year, current_date.month, 1)
        #     last_day_previous_month = first_day_current_month - timedelta(days=1)
        #     first_day_previous_month = first_day_current_month.replace(day=1)
        #     # Obter a data do último dia do mês retrasado
        #     last_day_two_months_ago = (first_day_previous_month - timedelta(days=1)).strftime('%Y-%m-%d').strip()
        #     first_day_two_months_ago = first_day_current_month.replace(month=current_month-2, day=1)
        #     last_day_previous_month = (first_day_two_months_ago.replace(month=current_month-1)
        #         - timedelta(days=1)).strftime('%Y-%m-%d').strip()
            
        #     self.initial_date = last_day_previous_month
        #     self.final_date = last_day_two_months_ago
        # except Exception as e: print(f'\x1b[36m>>> Erro em set_period(): {e}\x1b[0m')
        try:
            current_date = datetime.datetime.today()

            # Obter a data do último dia do mês passado
            first_day_current_month = date(current_date.year, current_date.month, 1)
            last_day_previous_month = first_day_current_month - timedelta(days=1)

            # Obter a data do último dia do mês retrasado
            if first_day_current_month.month > 2:
                first_day_two_months_ago = first_day_current_month.replace(month=current_date.month-2, day=1)
            else:
                # Se o mês atual for Janeiro ou Fevereiro, ajustar as datas manualmente
                last_day_previous_month = date(current_date.year - 1, 12, 31)
                first_day_two_months_ago = date(current_date.year - 1, 11, 30)

            self.initial_date = first_day_two_months_ago.strftime('%Y-%m-%d')
            self.final_date = last_day_previous_month.strftime('%Y-%m-%d')
            print(f'\x1b[36m>>> resultado de set_period(): \nData inicial: {self.initial_date}\nData Final: {self.final_date}\x1b[0m')

        except Exception as e:
            print(f'\x1b[36m>>> Erro em set_period(): {e}\x1b[0m')
    async def download_worksheet(self) -> None:
        """
        Baixar planilha
        """
        try:
            # Drop box
            self.driver.find_elements(
                By.CSS_SELECTOR,
                'button[class="ant-btn ant-btn-default ant-btn-sm ant-btn-icon-only ant-dropdown-trigger"]'
            )[-1].click()
            await asyncio.sleep(1)
            # Gerar relatório
            self.driver.find_element(
                By.CSS_SELECTOR,
                'div[class="ant-dropdown ant-dropdown-placement-bottomRight "]'
            ).find_elements(By.TAG_NAME, 'li')[-1].click()
            await asyncio.sleep(2)
            """
            Campos de datas
            """
            # Data inicial
            input_initial_date = self.driver.find_elements(By.CSS_SELECTOR, 'input[placeholder="Data inicial"]')[-1]
            input_initial_date.send_keys(Keys.BACKSPACE*10,self.initial_date, Keys.ENTER)
            await asyncio.sleep(1)
            input_initial_date.send_keys(Keys.BACKSPACE*10, self.initial_date, Keys.ENTER)
            await asyncio.sleep(1)
            # Data Final
            input_final_date = self.driver.find_elements(By.CSS_SELECTOR,
                'input[placeholder="Data final"]')[-1]
            input_final_date.send_keys(Keys.BACKSPACE)
            await asyncio.sleep(1)
            input_final_date.send_keys(Keys.BACKSPACE*10, self.final_date, Keys.ENTER)

            # print(f'>>> Data Inicial {self.initial_date}\nData Final {self.final_date}   ')

            # Remove cada arquivo CSV encontrado
            for file in os.listdir('./geracao'):
                if file.endswith('.csv'):
                    os.remove(os.path.join('./geracao', file))
            
            # Baixar csv (geração do cliente)
            self.driver.find_elements(By.CSS_SELECTOR,
                'a[class="ant-btn ant-btn-primary"]')[-1].click()
            # Fechar modal "Exportar dados de geração"
            self.driver.find_element(
                By.CSS_SELECTOR,
                'button[class="ant-modal-close"]'
            ).click()

            # Renomear o arquivo
            await asyncio.sleep(1)
            for file_name in os.listdir('./geracao'):
                if file_name.endswith(".csv"):
                    old_file_path = os.path.join('./geracao', file_name)
                    new_file_path = os.path.join('./geracao', "main.csv")
                    os.rename(old_file_path, new_file_path)
                    break
        except NoSuchElementException as er:
            print('>>> ', er)
            await self.download_worksheet()
            
    async def on_loaded(self) -> None:
        """"""
        # diretórios
        if not os.path.exists('./relatorios'):
            os.mkdir('./relatorios')
            print('>>> Diretório `relatórios` criado.')

        if not os.path.exists('./geracao'):
            os.mkdir('geracao')
            print('>>> Diretório `geracao` criado.')

        # Dataframes
        self.file_name = date.today().strftime('%m%Y')
        if not os.path.exists(f'./relatorios/{self.file_name}'):
            os.mkdir(f'./relatorios/{self.file_name}')

        if not os.path.exists(f'./relatorios/{self.file_name}/relatorio.xlsx'):
            df = pd.DataFrame()
            with pd.ExcelWriter(f'./relatorios/{self.file_name}/relatorio.xlsx') as writer:
                df.to_excel(writer, sheet_name='clientes', index=False)
            
    async def next_page(self) -> bool:
        """
        Passa para a próxima página
        """
        try:
            button_next_element = self.driver.find_element(By.CSS_SELECTOR, 'li[title="Próxima página"] > button')

            # button_next = self.driver.find_element(By.CSS_SELECTOR, 'li[title="Próxima página"] > button').click()
                    # Clique no botão usando JavaScript
            self.driver.execute_script("arguments[0].click();", button_next_element)
            await asyncio.sleep(2)

            print('>>> ➡️➡️')



        except Exception as e:pass
        """
        Verifica se o botão está habilitade
        Caso contrário chegou a última página
        """
        if not button_next_element.is_enabled():
            print('✖️➡️✖️')
            return False

        # button_next.click()
        # await asyncio.sleep(3)
        
        return True

    async def handle_worksheet(self, df, customer_name, customer_email, installation_date, performace, last_month) -> None:
        """
        manipular planilha relatorio
        """
        value_list = [
            customer_email,
            installation_date,
            performace,
            self.LIST_MONTHS[last_month - 1]
        ]
        dff = pd.read_csv('./geracao/main.csv')
        dff['Gerado(kWh)'] = dff.apply(lambda row: f"{row['Dia']},{row['Gerado(kWh)']}", axis=1)
        generation_list = dff['Gerado(kWh)'].tolist()
        value_list.extend(generation_list)
        
        try:
            df[f'{customer_name}'] = value_list
        except Exception as erro:
            print(f'>>> Erro, possível inconsistência de dados na planilha do cliente capturado\n {erro}')
    
    async def close_modal(self) -> None: 
        self.driver.find_element(By.CSS_SELECTOR, 'div[class="ant-modal-content"]')\
            .find_element(By.TAG_NAME, 'header').find_elements(By.CLASS_NAME, 'ant-col')[1]\
            .find_elements(By.TAG_NAME, 'button')[-1].click()
    
    async def get_customer_detail(self, page_number, update) -> dict:

        # Função para criar uma nova planilha Excel
        def criar_planilha(nome_arquivo):
                    if os.path.exists(nome_arquivo):
                        workbook = openpyxl.load_workbook(nome_arquivo)
                    else:
                        workbook = openpyxl.Workbook()

                        workbook.save(nome_arquivo)

        # Função para adicionar um novo usuário verificado à planilha
        def adicionar_usuario_verificado(nome_arquivo, usuario_verificado):
            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active

            # Adicionar os dados do usuário verificado como uma nova linha na planilha
            sheet.append(usuario_verificado)

            workbook.save(nome_arquivo)
    
        try:
            """"""
            await self.on_loaded()
            await self.set_period()
            usuarios_verificados = []
            data = {
                'success': False,
                'customers-without-email': [],
                'current-page': '',
                'verificateds':[]
            }

            try:
                iframe = self.driver.find_element(By.TAG_NAME, 'iframe')
                self.driver.switch_to.frame(iframe)
                self.driver.find_element(By.CSS_SELECTOR, 'li[title="1"]').click()
                
                
                await asyncio.sleep(1)
                                                                    # Criar a planilha se ela não existir
                criar_planilha(f'./relatorios/{self.file_name}/relatorio_detalhes.xlsx')
                ## ORDENAR POR GERAÇÃO MENSAL - MAIORES PRIMEIRO
                try:
                    # Clicar no elemento para ordenar por geração mensal
                    # Encontrar o elemento pelo XPath
                    ordenar_por_geracao = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, '#__next > div > div > div > div > div.ant-ribbon-wrapper > section > section > main > div:nth-child(4) > div > div > div > div > div > div > div.ant-table-header > table > thead > tr > th:nth-child(8) > div > span.ant-table-column-sorter.ant-table-column-sorter-full'))
                    )
                    # Executar o clique no elemento usando JavaScript
                    self.driver.execute_script("arguments[0].click();", ordenar_por_geracao)
                    print('>>> Clicou em ordenar por geração mensal')

                    await asyncio.sleep(1)

                    self.driver.execute_script("arguments[0].click();", ordenar_por_geracao)
                    print('>>> Clicou em ordenar por geração mensal novamente')
                    # Aguardar um tempo para que a página seja atualizada
                    await asyncio.sleep(5)

                    # Se necessário, clique novamente para inverter a ordem
                    # ordenar_por_geracao.click()
                    # print('>>> Clicou em ordenar por v geração mensal novamente (Maiores primeiro)')

                except Exception as e:
                    print('>>> Erro ao clicar em ordenar por geração mensal')
                    await asyncio.sleep(60)

                pagina = 1
                
                # Localizar o elemento usando o seletor CSS

                element = self.driver.find_element(By.XPATH, '//*[@id="__next"]/div/div/div/div/div[2]/section/section/main/div[1]/div[1]/div/div[2]/div[2]/h4/span')

                # Extrair o texto do atributo "textContent" do elemento
                element_text = element.text

                print("Total de usinas:", element_text)
                total_de_paginas = int(element_text) / 20 

                divisao = int(element_text)%20
                # print('Divisão: ', divisao)

                if divisao == 0:
                    ##paginas = paginas
                    total_de_paginas = total_de_paginas
                else:
                    if total_de_paginas < 1:
                        total_de_paginas = 1
                    else:
                        total_de_paginas = total_de_paginas + 1
                print(f'Total de páginas: {int(total_de_paginas)}')

                if page_number > 1:
                    for _ in range(1, page_number):
                        await self.next_page()
                        text_paginas =f' ➡️ 📑✅ Página {pagina} de {total_de_paginas:.0f}            🕐'
                        await update.message.reply_text(text_paginas)
                        pagina = pagina + 1



                last_month = ( datetime.datetime.now() - timedelta(days=30)).month
                current_year_last_month = ( datetime.datetime.now() - timedelta(days=30)).strftime("%Y-%m")
                customers_without_email = []
                
                contador = 0
                tem_mais_paginas = True
                usuarios_verificados = []
                retorno_paginas = True  ## se o botão False, aguarda o proximo loop para parar
                # while contador < 1:
                element_selector = '#__next > div > div > div > div > div.ant-ribbon-wrapper > section > section > main > div:nth-child(4) > div > div > div > div > ul > li.ant-pagination-item.ant-pagination-item-101 > a'
                # class="ant-pagination-item ant-pagination-item-101"
                usuario_descricao = ['Nome', 'Desempenho', 'Data da Instalação', 'E-mail', 'Mês', 'Status Geração', 'Telefone (TESTE)', 'Instação Antes da Data inicial', 'Telefone (CRM)']
                adicionar_usuario_verificado(f'./relatorios/{self.file_name}/relatorio_detalhes.xlsx',usuario_descricao)
                while(pagina <= total_de_paginas):
                    tem_mais_paginas = retorno_paginas
                    # contador += 1
                    await asyncio.sleep(1)
                    try:
                        # print('>>> Página atual: ', data['current-page'])

                        data['current-page'] = int(self.driver.find_elements(By.CLASS_NAME, 'ant-pagination-item-active')[0].text)
                    except NoSuchElementException as er:
                        print('>>> Erro ao tentar achar paginador.')
                        print('>>> ', er)
                    
                    try:
                        tem_mais_paginas = True
                        while (pagina <= total_de_paginas):
                            print(f'>>>> Página  {pagina}   de  {total_de_paginas:.0f}')
                            data['current-page'] = int(pagina)
                            try:
                                await asyncio.sleep(5)
                                tr_list = self.driver.find_elements(
                                By.CSS_SELECTOR,
                                'tr[class="ant-table-row ant-table-row-level-0"]'
                            )
                            except:
                                print('sos')       

                            # Ler planilha excel
                            try:
                                arquivo_excel = f'./relatorios/{self.file_name}/relatorio.xlsx'
                                df = pd.read_excel(arquivo_excel, sheet_name='clientes')
                            except:
                                # Se ocorrer um erro ao ler o arquivo, imprima uma mensagem de erro
                                # print(f'>>> [ERROR] Não foi possível ler a planilha excel')
                                
                                # Tente apagar o arquivo
                                try:
                                    os.remove(arquivo_excel)
                                    print('>>> Arquivo removido com sucesso.')
                                except Exception as e:
                                    # Se ocorrer um erro ao apagar o arquivo, imprima uma mensagem de erro
                                    pass
                                
                                # Agora crie um novo arquivo em branco
                                df = pd.DataFrame()
                                df.to_excel(arquivo_excel, sheet_name='clientes', index=False)
                                print('>>> Arquivo relatorio.xlsx criado.')
                            
                            try:
                                # adicionar indice para acompanhar as linhas da pagina 
                                # loop das 20 linhas por pagina
                                linha = 0
                                
                                for tr in tr_list:
                                    linha = linha + 1
                                    print(f'entrou no for -> linha: {linha}')
                                    try:
                                        customer = tr.text.split('\n')
                                        installation_date = str(datetime.datetime.strptime(
                                                customer[2][-10:], '%d/%m/%Y').date())
                                        customer_name = customer[0]
                                        def remover_acentos(texto):
                                            try:
                                                texto_sem_acentos = unidecode(texto)
                                                return texto_sem_acentos
                                            except Exception as e:
                                                raise ValueError("Erro ao remover acentos: {}".format(e))
                                        customer_name = remover_acentos(customer_name)
                                        performace = ''
                                        cabecalho = f'\n\x1b[32m>>>>> Página: \x1b[36m{pagina} \x1b[32m------ Linha: \x1b[36m{linha} \x1b[32m--------------------\n>>>>> Capturando dados do cliente:\x1b[0m  {customer_name}'
                                        print(cabecalho)

                                        # abrir modal (detalhes do cliente)
                                        try:
                                            self.driver.execute_script("arguments[0].click();",
                                                tr.find_element(By.CLASS_NAME, 'ant-space-item'))
                                            await asyncio.sleep(1)
                                        except Exception as e: print(e)
                                        """Abre nova aba, detalhes da usina"""
                                        try:
                                            self.driver.find_element(By.CSS_SELECTOR, 'div[class="ant-modal-content"]')\
                                                .find_element(By.TAG_NAME, 'header').find_elements(By.TAG_NAME, 'a')[-1]\
                                                .click()
                                        except:
                                            # customers_without_email.append({'name' : customer_name})
                                            print(f'\x1b[33m>>>>> Cliente não cadastrado para a usina \x1b[0m{customer_name}')
                                        # exit
                                            try:
                                                customer_email = 'email@indisponivel'

                                                # await self.close_modal()
                                            except:
                                                continue
                                        
                                        """Vai para a aba de detalhes da usina; Capturar informações"""
                                        try:
                                            await asyncio.sleep(1)
                                            self.driver.switch_to.window(self.driver.window_handles[1])
                                            {"msg":"Recurso não encontrado para este integrador"}
                                            try:
                                                customer_email = self.driver.find_elements(By.CLASS_NAME, 'inner')[0].text
                                                email_text = customer_email.split('\n')
                                                for item in email_text:
                                                    if '@' in item:  # Check if the item contains '@' (indicative of an email)
                                                        customer_email = item.strip()  # Remove any leading/trailing spaces
                                                        break  # Stop processing after finding the first email-like item
                                            except:
                                                customer_email = 'email@indisponivel'

                                            print(f'\x1b[32m>>>>> email : \x1b[0m{customer_email}')
                                                                    
                                            self.driver.close()
                                            self.driver.switch_to.window(self.driver.window_handles[0])
                                            iframe = self.driver.find_element(By.TAG_NAME, 'iframe')
                                            self.driver.switch_to.frame(iframe)
                                            await asyncio.sleep(1)

                                        except:
                                            pass

                                        try:
                                            try:
                                                self.driver.find_element(
                                                    By.CSS_SELECTOR,
                                                    'div[title="Mês"]'
                                                ).click()
                                                await asyncio.sleep(1)
                                            except:
                                                button_mes = self.driver_find_element(By.XPATH, '/html/body/div[2]/div/div[2]/div/div[2]/div/section/div/div[1]/section/div[1]/div[1]/div/div/label[3]/div')
                                                button_mes.click()

                                            input_month = self.driver.find_element(
                                                By.CLASS_NAME, 'ant-picker-input'
                                            ).find_element(By.TAG_NAME, 'input')
                                            input_month.send_keys(Keys.BACKSPACE)
                                            await asyncio.sleep(1)
                                            input_month.send_keys(
                                                Keys.BACKSPACE*10, 
                                                current_year_last_month,
                                                Keys.ENTER
                                            )
                                            await asyncio.sleep(3)
                                        except NoSuchElementException as er:
                                            print('>>> [ERROR] ao manipular data do desempenho mensal.')
                                            # print(er)
                                            # exit
                                            
                                            data['success'] = False
                                            return data
                                            
                                        try:
                                            performace = self.driver.find_elements(
                                                By.CSS_SELECTOR,
                                                'span[class="block font-semibold"]'
                                            )[-1].text
                                        except NoSuchElementException as er:
                                            print('>>> [ERROR] ao capturar desempenho percentual.')
                                            print(er)
                                            # exit
                                            

                                            return data
                                        

                                        try:
                                            performace = performace.replace('.','')
                                            desempenho = performace.replace('%','')
                                            desempenho = desempenho.replace(',','.')
                                        except: 
                                            desempenho = performace
                                        # desempenho = 0
                                        try:
                                            def valid_installation(installation_date,data_inicial):
                                                try:
                                                    valid = installation_date < data_inicial
                                                    # print(valid)
                                                    if valid == True:
                                                        intalacao = 'before'
                                                        return True
                                                    else:
                                                        print("Instalação da usina foi depois da data inicial")
                                                        intalacao = 'after'
                                                        return False
                                                except Exception as e: print(e)
                                            # await self.set_period()
                                            data_inicial = self.initial_date
                                            instalacao = valid_installation(installation_date,data_inicial)
                                        except Exception as e: print(f'>>>[465]: ', e)
                                        # print(f'>>> Desempenho: {desempenho} %\nData da Instalação: {installation_date}\n{data_inicial}')
                                        
                                        # PLANILHA DE ERROS
                                        try:
                                                if float(desempenho) >= 90:
                                                    status = 'Geração Excelente'
                                                elif float(desempenho) >= 72 and float(desempenho) < 90:
                                                    status = 'Geração Ok'
                                                elif float(desempenho)>0 and float(desempenho)<72:
                                                    status = 'Geração abaixo do esperado'
                                                elif float(desempenho)== 0:
                                                    status = 'Geração Zerada!'
                                                else:
                                                    status = desempenho
                                                try:                        # cliente, desempenho, data de instalação, email, mês referência, status
                                                    
                                                    telefone_do_cliente = '17991028000'
                                                    
                                                    def obter_telefones(nome, email):
                                                        try:
                                                            # Carregue a planilha usuarios_crm.xlsx em um DataFrame
                                                            df = pd.read_excel('usuarios_crm.xlsx')

                                                            # Verifique se o email existe na planilha
                                                            if 'email' in df.columns:
                                                                email_match = df[df['email'] == email]
                                                                if not email_match.empty:
                                                                    telefones = email_match['telefones'].tolist()[0]  # Pegue a primeira correspondência
                                                                    return extrair_numeros(telefones)

                                                            # Verifique se o nome existe na planilha
                                                            if 'nome' in df.columns:
                                                                nome_match = df[df['nome'] == nome]
                                                                if not nome_match.empty:
                                                                    telefones = nome_match['telefones'].tolist()[0]  # Pegue a primeira correspondência
                                                                    return extrair_numeros(telefones)

                                                            try:
                                                                df.close()  # Feche o DataFrame
                                                            except:pass
                                                            return 'Cliente não localizado no CRM'
                                                        except Exception as e:
                                                            try:
                                                                df.close()  # Feche o DataFrame
                                                            except:pass
                                                    def extrair_numeros(telefones):
                                                        # Use expressões regulares para extrair todos os números dos telefones
                                                        numeros = re.findall(r'\d{2,}', telefones)
                                                        # print(f'numeros: {numeros}')
                                                        try:
                                                            numeros = f'{"".join(numeros)}'
                                                            # print('Numeros = ', numeros)d
                                                        except Exception as e: print(e)
                                                        if numeros:
                                                            return numeros
                                                        else:
                                                            return None
                                                    
                                                    telefones = obter_telefones(customer_name, customer_email)

                                                    if telefones:
                                                        print("\x1b[32m>>>>> Telefones:\x1b[0m" ,telefones)
                                                    else:
                                                        telefones = 'Número não cadastrado no CRM'
                                                        print("\x1b[33m>>>>> Número não cadastrado no CRM\x1b[0m")

                                                        # print("Cliente não encontrado na planilha.")

                                                    usuario_verificado = [customer_name, desempenho, installation_date, customer_email,last_month , status, telefones, instalacao, telefones]
           
                                                    # Adicionar o usuário verificado à planilha
                                                    adicionar_usuario_verificado(f'./relatorios/{self.file_name}/relatorio_detalhes.xlsx', usuario_verificado)
                                                    
                                                except Exception as e: print(':::>>>>>', e)
                                        except Exception as erro: print(f'ERRO ERRO', erro)
                                        
                                        ## comparar installation_date e data_inicial 
                                        try:
                                            if float(desempenho) == 0:
                                                print('\x1b[31m>>> Desempenho Zerado !!!\x1b[0m')
                                            if float(desempenho) > 0 and float(desempenho) < 72:
                                                print(f'\x1b[33m>>>>> Desempenho abaixo do esperado.    {desempenho} %\x1b[0m')
                                            
                                                
                                            
                                            if float(desempenho) >= 1 and instalacao == True :
                                                ## DOWNLOAD DA GERAÇÃO MENSAL 
                                                try:
                                                    await self.download_worksheet()
                                                    # print('>>>> DOWNLOAD DO HISTÓRICO DE GERAÇÃO OK!')
                                                except:
                                                    print(f'>>> [ERROR] ao baixar planilha do usuario {customer_name}')
                                                    # exit
                                                    # 
                                                    # return data
                                                
                                                # print(f'{customer_name} {customer_email} {installation_date} {performace} {last_month}')
                                                ## MANIPULAÇÃO DO ARQUIVO CSV 
                                                try:
                                                    await self.handle_worksheet(df, customer_name, customer_email,
                                                        installation_date, performace, last_month)
                                                    # print('>>> HANDLE_WORKSHEET OK !')
                                                    # print(df, customer_name, customer_email,
                                                        # installation_date, performace, last_month)
                                                    await asyncio.sleep(1)
                                                except Exception as e: print(e)

                                                ## FECHAR POP UP (??)
                                                try:
                                                    await self.close_modal()
                                                    await asyncio.sleep(1.5)
                                                except NoSuchElementException as er:
                                                    print('>>> [ERROR] ao fechar o modal')
                                                    print(er)
                                                    # exit
                                                    data['success'] = False
                                                    # return data

                                                print(f'\x1b[32m>>>>> Desempenho Ok.    {desempenho} %\x1b[0m')
                                                
                                                # print('Usuarios Verificados')
                                                # print(usuarios_verificados)
                                                # print('')  
                                                print(f'\x1b[32m✅ {customer_name} finalizado!\n----------------------------------------------------\x1b[0m')
                                        except Exception as e: print('', end="")

                                    except Exception as e: 
                                        print('>>> erro[504]: ')
                                        data['success'] = False
                                 
                                    with pd.ExcelWriter(f'./relatorios/{self.file_name}/relatorio.xlsx') as writer:
                                        df.to_excel(writer, sheet_name='clientes', index=False)
                                # ================================ 
                                # print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')
                                # await asyncio.sleep(35)

                                # pular pagina no final da execução
                                retorno_paginas =  await self.next_page()
                                if retorno_paginas == True:
                                    text_paginas =f'  📑✅ Página {pagina} de {total_de_paginas:.0f}          🕐'
                                    await update.message.reply_text(text_paginas)
                                    pagina = pagina + 1
                                else:
                                    print('Ultima pagina')
                                    pagina = pagina + 1
                                    break




                                # print(f'<<<<<<<<<<<<<<<<<<<<<<<<<<<<<')
                            except Exception as e: 
                                print('>>>>546: ', e)

                                data['success'] = False
                                return data
                        # print('Saiu do segundo while' )
                        # print(f'pagina = {pagina}    total_de_paginas = {total_de_paginas}')

                    except Exception as er:
                        print(f'>>> [Error543]: {er}')
                        # print('>>> data:', data)
                        # exit 
                        
                        data['success'] = False

                        return data
                
            except Exception as e: print('<< Get_customer_details: ', e)
            data_status = data['success']
            # print(f'>>> data= {data_status}<<<< ')
            if data_status == 'False':
                return data
            else:
                data['success'] = True
                try:
                    await self.driver_finish()
                finally:
                    pass

                print(f'💾  Retornar data ')
                return data
        except Exception as er:
            print(er)
            return data

    async def get_customers(self) -> dict:
        """Pega informações dos clientes no site Solarz, como
        nome, data de instalação e valor do desempenho"""
        
        data = {'path': True}
        customer_list = []

        iframe = self.driver.find_element(By.TAG_NAME, 'iframe')

        self.driver.switch_to.frame(iframe)
        self.driver.find_element(By.CSS_SELECTOR, 'li[title="1"]').click()
        await asyncio.sleep(1)
        
        contador = 0
        # for pagina in paginas
        while contador < 1:
            contador += 1
            try:
                ## extrai informações das 20 linhas da pagina
                tr_list = self.driver.find_elements(
                    By.CSS_SELECTOR,
                    'tr[class="ant-table-row ant-table-row-level-0"]'
                )

                # loop de execução das 20 linhas (for linha in linhas)
                for tr in tr_list:
                    try:
                        customer = tr.text.split('\n')
                        if len(customer) == 7:
                            installation_date = str(datetime.datetime.strptime(
                                customer[2][-10:], '%d/%m/%Y').date())
                            customer_list.append({
                                'name': customer[0],
                                'kwp': customer[1],
                                'installation_date': installation_date,
                                'performace': customer[3]
                            })

                    except NoSuchElementException as er: print(er)

                ## apos as 20 linhas (tr_list) clicar em proxima pagina
                try:
                    button_next = await self.next_page()

                    # print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>.')
                    ## se o botão de proxima pagina retornar False, para a execução
                    if not button_next.is_enabled():
                       break
                    
                    await asyncio.sleep(1)

                except NoSuchElementException as er:
                     print(f'>>> [Error609]: {er}')
                
            except NoSuchElementException as er:
                print(f'>>> [Error612]: {er}')

        if not os.path.exists('executivos'):
            os.makedirs('executivos')
            df = pd.DataFrame(columns=['Nome do cliente', 'Nome alternativo', 'Executivo', 'Criação da usina'])
            with pd.ExcelWriter('executivos/MONITORAMENTO.xlsx') as writer:
                df.to_excel(writer, sheet_name='MONITORAMENTO', index=False)
            data['path'] = False
            print(f'>>> { datetime.datetime.now()} Diretório executivos criado!')

        # Manipular planilha excel
        try:
            df = pd.read_excel('executivos/MONITORAMENTO.xlsx', sheet_name='MONITORAMENTO')
        except:
            df = pd.DataFrame(columns=['Nome do cliente', 'Nome alternativo', 'Executivo', 'Criação da usina'])
            with pd.ExcelWriter('executivos/MONITORAMENTO.xlsx') as writer:
                df.to_excel(writer, sheet_name='MONITORAMENTO', index=False)
            data['path'] = False
            print(f'>>> { datetime.datetime.now()} Arquivo excel MONITORAMENTO.xlsx criado!')
        
        today = date.today().strftime("%d/%m/%Y")
        
        for customer in customer_list:
            if not df['Nome do cliente'].isin([customer['name']]).any():
                installation_date = datetime.strptime(
                    customer['installation_date'], "%Y-%m-%d")
                df.loc[len(df)] = {
                    'Nome do cliente': customer['name'],
                    'Criação da usina': installation_date.strftime("%d/%m/%Y")
                }

            df.loc[df['Nome do cliente'] == customer['name'],
                   today] = customer['performace']

        # Salvar xlsx
        with pd.ExcelWriter('executivos/MONITORAMENTO.xlsx') as writer:
            df.to_excel(writer, sheet_name='MONITORAMENTO', index=False)
        
        
        return data

    async def apply_styling(self) -> None:
        """Aplica estilização de cores na tabela master, somente!"""
        file_path = 'master.xlsx'
        
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()
            
        # Remove a aba "Sheet"
        if "Sheet" in workbook.sheetnames:
            worksheet = workbook["Sheet"]
            workbook.remove(worksheet)

        red_fill = PatternFill(start_color='DD1C1A',
                            end_color='DD1C1A', fill_type='solid')
        orange_fill = PatternFill(
            start_color='E2711D', end_color='E2711D', fill_type='solid')
        yellow_fill = PatternFill(
            start_color='F0C808', end_color='F0C808', fill_type='solid')
        green_fill = PatternFill(start_color='53A548',
                                end_color='53A548', fill_type='solid')

        # Percorre todas as abas do arquivo Excel
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Seleciona todas as células a partir da terceira coluna, ignorando a primeira linha
            for row in worksheet.iter_rows(min_row=2, min_col=3):
                for cell in row[2:]:
                    if cell.value is None:
                        continue
                    cell_value = int(cell.value)
                    if cell_value == 0:
                        cell.fill = red_fill
                    elif cell_value > 0 and cell_value <= 60:
                        cell.fill = orange_fill
                    elif cell_value > 60 and cell_value < 90:
                        cell.fill = yellow_fill
                    elif cell_value >= 90:
                        cell.fill = green_fill

        # Salva o arquivo Excel com as mudanças
        workbook.save(file_path)

    async def set_executive_worksheet(self, executive_name:str) -> None:
        """"""
        file_name = executive_name.upper()
        
        df_monitoramento = pd.read_excel('executivos/MONITORAMENTO.xlsx', sheet_name='MONITORAMENTO')

        try:
            df = pd.read_excel(f'executivos/{file_name}.xlsx', sheet_name=file_name)
        except FileNotFoundError:
            df = pd.DataFrame()
            df.to_excel(f'executivos/{file_name}.xlsx', sheet_name=file_name, index=False)
            
        df_filtered = df_monitoramento.loc[df_monitoramento['Executivo'] == executive_name]
        
        df_filtered.to_excel(f'executivos/{file_name}.xlsx',
            sheet_name=file_name, index=False)

    async def get_proposal(self) -> dict:
        """"""
        data = {
            'success': False,
            'error': '',
            'seller_without_executive': [],
            'seller_not_found': [],
            'clients_not_found': []
        }
        df_customer = pd.read_excel('executivos/MONITORAMENTO.xlsx', sheet_name='MONITORAMENTO')
        
        if os.path.exists('executivos/MONITORAMENTO.xlsx'):
            customer_list = pd.read_excel(
                'executivos/MONITORAMENTO.xlsx',
                usecols=['Nome do cliente', 'Nome alternativo',
                        'Executivo', 'Criação da usina']).to_dict('records')
        else:
            print('Arquivo `MONITORAMENTO.xlsx` não encontrado!')
            data['error'] = 'Arquivo `MONITORAMENTO.xlsx` não encontrado!'
            return data
        
        if os.path.exists('executivos/AUXILIAR.xlsx'):
            executive_list = pd.read_excel(
                'executivos/AUXILIAR.xlsx',
                usecols=['Franqueado/vendedor responsável', 'Executivo de engenharia']).to_dict('records')
        else:
            print('Arquivo `AUXILIAR.xlsx` não encontrado!')
            data['error'] = 'Arquivo `AUXILIAR.xlsx` não encontrado!'
            return data
        
        # Ir para página de propostas
        self.driver.get(self.URLS_SITE['proposals'])
        """Definir tamanho do navegador para todos
        os elementos serem renderizados na tela"""
        self.driver.set_window_size(1920, 1080)
        await asyncio.sleep(2)
        
        filter_proposal = self.driver.find_element(
            By.CSS_SELECTOR,
            'input[placeholder="Código Proposta ou Nome Cliente"]'
        )

        for customer in customer_list:
            name_to_search = customer['Nome alternativo']
            if type(name_to_search) == float:
                name_to_search = customer['Nome do cliente']
            filter_proposal.clear()
            filter_proposal.send_keys(name_to_search, Keys.ENTER)
            await asyncio.sleep(4)

            try:
                card = self.driver.find_element(
                    By.CSS_SELECTOR, 'article[class="card"]')
            except NoSuchElementException as er:
                print(f'>>> Cliente {customer["Nome do cliente"]} não localizado')
                data['clients_not_found'].append(customer["Nome do cliente"])
                continue
            

            # Seleciona o card
            self.driver.execute_script("arguments[0].click();", card)
            await asyncio.sleep(1)
            # Ir para a aba de detalhes do cliente que foi aberta
            self.driver.switch_to.window(self.driver.window_handles[1])
            await asyncio.sleep(1)
            # Pega o nome no select
            seller = Select(self.driver.find_element(
                By.NAME, 'user_id')).first_selected_option.text
            seller = str(seller).strip()
            
            # Fecha a aba de detalhes aberta
            self.driver.close()
            # Ir para aba da tela de proposta
            self.driver.switch_to.window(self.driver.window_handles[0])

            # Localizar o executivo pelo vendedor/franqueado
            executive = next(
                (e for e in executive_list if e['Franqueado/vendedor responsável'] == seller), None)

            executive_aux = ''

            if not executive or 'Executivo' in executive:
                print(f'>>> Vendedor/franqueado: {seller} não localizado.')
                if not seller in data['seller_not_found']:
                    data['seller_not_found'].append(seller)
            elif type(executive['Executivo de engenharia']) == float:
                print(f'>>> Vendedor/franqueado: {seller} não poussuí executivo vinculado.')
                if not executive['Franqueado/vendedor responsável'] in data['seller_without_executive']:
                    data['seller_without_executive'].append(executive['Franqueado/vendedor responsável'])
            else:
                executive_aux = executive['Executivo de engenharia']
            
            # Atualizar o vinculo na planilha de monitoramento
            df_customer.loc[
                df_customer['Nome do cliente'] == customer['Nome do cliente'],
                'Executivo'] = executive_aux
            
            # Salvar alteração na planilha master
            with pd.ExcelWriter('executivos/MONITORAMENTO.xlsx') as writer:
                df_customer.to_excel(writer, sheet_name='MONITORAMENTO', index=False)
            
            if executive_aux:
                print(f'>>> Vinculando cliente: {customer["Nome do cliente"]} ao executivo: {executive["Executivo de engenharia"]}')
                # Criar/atualiza a planilha de executivo
                self.set_executive_worksheet(executive_name=executive_aux)
        
        data['success'] = True
        return data
    

    async def driver_finish(self) -> None:
        self.driver.close()
        self.driver.quit()

    async def create_customer_list(self, dataframe, filtered_data) -> list:
        customer_list = []

        for _, linha in filtered_data.iterrows():
            client = {'name': linha['Nome do cliente'],
                      'created': linha['Criação da usina']}
            for i in range(2, len(dataframe.columns)):
                client[dataframe.columns[i]] = linha.iloc[i]

            customer_list.append(client)

        return customer_list

    async def filter_bad_performance(self, df) -> list:
        """"""
        try:
            return await self.create_customer_list(
                df,
                df[(df.iloc[:, -1] == 0) & (df.iloc[:, -2] > 0)]
            )
        except:
            return []

    async def filter_good_performance(self, df) -> list:
        """Filtra por usísinas que tiveram ótima performace naquele dia"""
        try:
            return await self.create_customer_list(
                df,
                df[(df.iloc[:, -1] > df.iloc[:, -2]) & (df.iloc[:, -2] == 0)]
            )
        except:
            return []

    async def filter_last_three_days_performance(self, df) -> list:
        """Filtra usínas que tiveram uma péssima perfomance nos
        últimos 3 dias"""
        try:
            column_before_dates = df.columns.get_loc('Criação da usina')
            if not len(df.columns) >= column_before_dates + 4:
                print('Não foi possível aplicar filtro de 3 dias, devido a quantidade de dias presentes na tabela.')
                return []
        
            return await self.create_customer_list(
                df,
                df[df.iloc[:, -3:].apply(lambda x: all(x == 0), axis=1)]
            )
        except:
            return []

    async def filter_last_seven_days_performance(self, df) -> list:
        """Filtra pelas as usínas que tiverem uma média abaixo
        de 75% nos últimos 7 dias"""
        try:
            column_before_dates = df.columns.get_loc('Criação da usina')
            if not len(df.columns) >= column_before_dates + 8:
                print('Não foi possível aplicar filtro de 7 dias, devido a quantidade de dias presentes na tabela.')
                return []
            
            average = df.iloc[:, -7:].mean(axis=1)
            return await self.create_customer_list(
                    df,
                    df[(average < 75) & (average != 0)]
            )
        except:
            return []

    async def merge_dataframe(self) -> None:
        """Junta todos os dataframes criando o arquivo master.xlsx"""
        new_book = Workbook()

        list_worksheets = os.listdir('./executivos')

        # Percorrendo os arquivos existentes na pasta
        for filename in list_worksheets:
            if not filename.endswith('.xlsx'):
                continue

            # Carregando o arquivo existente
            book = load_workbook(f'executivos/{filename}')

            # Copiando cada aba para o novo arquivo
            for sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                new_sheet = new_book.create_sheet(sheet_name)
                for row in sheet:
                    for cell in row:
                        new_sheet[cell.coordinate].value = cell.value

        # Salvando o novo arquivo
        new_book.save('master.xlsx')
        await self.apply_styling()

    async def anilizy_clients(self) -> dict:
        """"""
        data = {
            'bad_performance': [],
            'good_performance': [],
            'last_three_days': [],
            'last_seven_days': [],
        }

        df = pd.read_excel('executivos/MONITORAMENTO.xlsx')
        data['bad_performance'] = await self.filter_bad_performance(df)
        data['good_performance'] = await self.filter_good_performance(df)
        data['last_three_days'] = await self.filter_last_three_days_performance(df)
        data['last_seven_days'] = await self.filter_last_seven_days_performance(df)

        return data


    async def monitoramento(self, user_choice, update) -> None:
        def erro(erro):
            hora_atual = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log = f'>>> [{hora_atual}] [ERROR] {erro}'
            print(log)

        query = update.callback_query


        try:
            """Realiza login no site"""
            self.driver.get(self.URLS_SITE['login_solarz'])
            await asyncio.sleep(2)
        except Exception as e: erro(e)
        if user_choice == 'engenharia':
            usuario = 'engenharia@kinsolenergia.com.br'
            senha = '@Kinsol21'
        else:
            usuario = 'juliag.kinsol@gmail.com'
            senha = 'Kinsol21'
        try:
            self.driver.find_element(
                By.ID, 'username').send_keys(usuario)
            self.driver.find_element(By.ID, 'password').send_keys(
                senha)
            self.driver.find_element(
                By.CSS_SELECTOR, 'input[value="Entrar"]').click()
            await asyncio.sleep(5)
        except Exception as er:
            print('>>> [ERROR]: Erro ao localizar elemento')
            print(f'>>> [ERROR]83: {er}')

        id = 0
        #VERIFICAR SE O LOGIN FOI REALIZADO COM SUCESSO
        logar = self.driver.current_url
        if logar == 'https://app.solarz.com.br/integrador/dashboard/v3':
                logado = 'sim'
                print(f'Sucesso ao logar! Usuário: {usuario}\n')
        
        else:
                logado = 'não'
                print('Erro ao realizar ou verificar login. Repita o processo.\n')
            
        if logado == 'sim':    
            # Muda o foco para o iframe 
            iframe = self.driver.find_element("xpath",'//*[@id="menu-tailwind"]/iframe')
            self.driver.switch_to.frame(iframe) 
            # Ordenar em ordem alfabética
            ordenar_alfabetico = WebDriverWait(self.driver, 20).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '#__next > div > div > div > div > div.ant-ribbon-wrapper > section > section > main > div:nth-child(4) > div > div > div > div > div > div > div.ant-table-header > table > thead > tr > th:nth-child(2) > div > span.ant-table-column-title > div'))
                        )
            # Executar o clique no elemento usando JavaScript
            self.driver.execute_script("arguments[0].click();", ordenar_alfabetico)
            print('>>> Clicou em ordenar alfabeticamente...')

            # Ver quantidade de páginas

            # Encontra o elemento pelo nome da classe e extrai o texto
            total_paginas = self.driver.find_element(By.CLASS_NAME, "ant-pagination-total-text").text
            print('Total de páginas: ',total_paginas)

            # Imprime o texto extraído
            # print(total_paginas)
            paginas = total_paginas.split(" ")
            pg_inicial = 1
            pg_final = int(int(paginas[1])/20)

            # Criar arquivo no Excel
            wb = Workbook()
            ws = wb.active
            from datetime import datetime, timedelta
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['D'].width = 35
            ws.cell(row=1,column=1).font = Font(size = 12, bold = True)
            ws.cell(row=1,column=2).font = Font(size = 12, bold = True)
            ws.cell(row=1,column=3).font = Font(size = 12, bold = True)
            ws.cell(row=1,column=4).font = Font(size = 12, bold = True)
            ws['A1'] = 'ID'
            ws['B1'] = 'USINA'
            ws['C1'] = 'GERAÇÃO DIA ANTERIOR'
            # ws['D1'] = 'STATUS'
            # Obter a data do dia anterior
            data_dia_anterior = datetime.now() - timedelta(days=1)
            dia_anterior = data_dia_anterior.strftime('%d/%m/%Y')

            ws['C1'] = f'{dia_anterior}'

            #formatação condicional
            try:
                    #icon set in column K
                    first = FormatObject(type='num', val=0)
                    second = FormatObject(type='num', val=1)
                    third = FormatObject(type='num', val=69)
                    four = FormatObject(type='num',val = 90 )

                    iconset = IconSet(iconSet='4TrafficLights', cfvo=[first, second, third, four], showValue=None, reverse=None)
                    rule = Rule(type='iconSet', iconSet=iconset)
                    ws.conditional_formatting.add("C2:C20000", rule)
            except Exception as e: print(e)
            
            ws.auto_filter.ref = "A1:D20000"


            for j in range(pg_inicial, pg_final+2):
                    def clicar(css):
                                    try:
                                        self.driver.execute_script(
                                            "arguments[0].click();",
                                            self.driver.find_element(By.CSS_SELECTOR, css)
                                        )
                                    except: pass
                                
                    print(f'******** página {j} ********')
                    await asyncio.sleep(2)
                    # rodar as 20 linhas
                    mensagem = ''
                    for i in range(2,22):
                    #    print(f'linha {i}')
                        try:
                            xpath_usina = f'//*[@id="__next"]/div/div/div/div/div[2]/section/section/main/div[4]/div/div/div/div/div/div/div[2]/table/tbody/tr[{i}]/td[2]/div/div[2]/div[1]/div/div/div[1]/h1'
                            css_usina = f'#__next > div > div > div > div > div.ant-ribbon-wrapper > section > section > main > div:nth-child(4) > div > div > div > div > div > div > div.ant-table-body > table > tbody > tr:nth-child({i}) > td:nth-child(2) > div > div.ant-col.gutter-row > div:nth-child(1) > div > div > div:nth-child(1) > h1'
                            # usina = driver.find_element("xpath", f'//*[@id="__next"]/div/div/div/div/div[2]/section/section/main/div[4]/div/div/div/div/div/div/div[2]/table/tbody/tr[{i}]/td[2]/div/div[2]/div[1]/div/div/div[1]/h1')

                            usina = self.driver.find_element(By.CSS_SELECTOR, css_usina).text


                            # print('>> Nome da usina...', usina)
                            # print(usina)
                            
                            xpath_desempenho = f'//*[@id="__next"]/div/div/div/div/div[2]/section/section/main/div[4]/div/div/div/div/div/div/div[2]/table/tbody/tr[{i}]/td[3]/div/div/span'
                            css_desempenho = f'#__next > div > div > div > div > div.ant-ribbon-wrapper > section > section > main > div:nth-child(4) > div > div > div > div > div > div > div.ant-table-body > table > tbody > tr:nth-child({i}) > td:nth-child(6) > div > div > span'
                            
                            # desempenho = driver.find_element("xpath", f'//*[@id="__next"]/div/div/div/div/div[2]/section/section/main/div[4]/div/div/div/div/div/div/div[2]/table/tbody/tr[{i}]/td[3]/div/div/span').text
                            desempenho = self.driver.find_element(By.CSS_SELECTOR, css_desempenho).text
                            print(f'\x1b[33m{desempenho}%        {usina}     \x1b[0m')
                            mensagem += f'{desempenho} %       {usina}\n'
                            await asyncio.sleep(0.1)
                            

                            # async def abrir_extrair_modal(i):
                                
                            #     # clicar_na_usina
                            #     try:
                            #         css_usina_desc = f'#__next > div > div > div > div > div.ant-ribbon-wrapper > section > section > main > div:nth-child(4) > div > div > div > div > div > div > div.ant-table-body > table > tbody > tr:nth-child({i}) > td.ant-table-cell.ant-table-column-sort > div > div.ant-col.gutter-row > div:nth-child(1) > div > div > div:nth-child(1) > h1'
                            #         # self.driver.find_element(By.CSS_SELECTOR, css_usina_desc).click()
                            #         clicar(css_usina_desc)
                            #     except Exception as e: print(f'Erro ao clicar na usina: {e}')    
                            #     await asyncio.sleep(1)
                            #     # clicar no botão dia
                            #     try:
                            #             self.driver.find_element(
                            #                         By.CSS_SELECTOR,
                            #                         'div[title="Dia"]'
                            #                     ).click()
                            #             await asyncio.sleep(5)
                            #     except Exception as e: print(f'Erro ao clicar no botão Dia: {e}')    
                            #     # extrair dados de desempenho da usina 
                            #     try:
                            #             desempenho_encontrado = False
                                        
                            #             try:

                            #                 # Agora, você pode listar todos os elementos da página
                            #                 elementos = self.driver.find_elements_by_class_name("text-center")  # Isso obtém todos os elementos da página
                            #                 # Itere pelos elementos e procure a classe "text-center"
                            #                 desempenho = 0
                            #                 for elemento in elementos:
                            #                     class_text = elemento.get_attribute('class')
                            #                     if 'text-center' in class_text.split():
                            #                         # print("Tag:", elemento.tag_name)
                            #                         # print("Texto:", elemento.text)

                            #                         # Verifique se o texto do elemento contém "Desempenho"
                            #                         if "Desempenho" in elemento.text:
                            #                             texto_do_elemento = elemento.text.split('Desempenho')
                            #                             desempenho = texto_do_elemento[1].strip()
                            #                             desempenho = desempenho.replace(',','.')
                            #                             desempenho_encontrado = True
                            #                         else: 
                            #                             desempenho_encontrado = False                                                
                                            
                            #                 try:
                            #                     print(f'\x1b[32m {usina}  {desempenho}  \x1b[0m')
                            #                     print("\x1b[36m-\x1b[0m" * 25)
                            #                     print('\n')
                            #                     print("\x1b[36m-\x1b[0m" * 25)


                            #                 except Exception as e: print(f'\x1b[32m Erro >>> {e} \x1b[0m')

                            #             except NoSuchElementException as er:
                            #                 print('>>> [ERROR] ao capturar desempenho percentual.')
                            #                 print(er)
                            #     except Exception as e: print(f'Erro ao extrair desempenho: {e}')    
                            #     await asyncio.sleep(1)
                                
                            #     # clicar no X 
                            #     try:
                            #         css_X = 'body > div:nth-child(7) > div > div.ant-modal-wrap > div > div.ant-modal-content > div > header > div > div:nth-child(2) > div > div:nth-child(3) > button'
                            #         clicar(css_X)
                            #     except Exception as e: print(f'Erro ao fechar Modal: {e}')    
                                
                            # await abrir_extrair_modal(i)
                            id +=1
                            try:
                                inserir = [id, usina,int(desempenho)]
                                ws.append(inserir)
                            except Exception as e:print(e)
                        except Exception as erro:
                            pass
                        
                            
                    # clicar em próxima página
                    try:
                        
                        botao_prox = self.driver.find_element(By.CLASS_NAME,"ant-pagination-next").click()
                    except:

                        try:
                            self.driver.execute_script(
                                "arguments[0].click();",
                                self.driver.find_element(By.CLASS_NAME, "ant-pagination-next")
                            )
                        except Exception as e: print(f'Erro ao pular página: {e}')
                                
                    try:
                        await query.message.reply_text(mensagem)

                    except Exception as e: print(e)
            
            # Salvar arquivo
        
            # centralizar colunas excel
            try:
                    k=1
                    while k < 3000:
                            cellA = ws.cell(row=i, column= 1)
                            cellA.alignment = Alignment(horizontal="center")
                            
                            cellC = ws.cell(row=i, column= 3)
                            cellC.alignment = Alignment(horizontal="center")

                            cellD = ws.cell(row=i, column= 4)
                            cellD.alignment = Alignment(horizontal="center")
                            k+=1
                    
                    print("Colunas centralizadas")
            except Exception as e: print(e)    
            
            #salvar arquivo excel 
            try:
                    wb.save("monitoramento.xlsx")
                    print("Arquivo Salvo")
            except:
                    #arquivo ja está aberto e não deixa salvar por cima
                    try:
                            autoit.win_activate('monitoramento - Excel')
                            autoit.win_close('monitoramento - Excel')
                            pyautogui.press('right')
                            pyautogui.press('ENTER')
                            await asyncio.sleep(1)
                            wb.save("monitoramento.xlsx")
                    except:
                            try:
                                autoit.win_activate('monitoramento.xlsx - Excel')
                                autoit.win_close('monitoramento.xlsx - Excel')
                                pyautogui.press('right')
                                pyautogui.press('ENTER')
                                await asyncio.sleep(1)
                                wb.save("monitoramento.xlsx")
                            except Exception as e: print(e)
            return True
