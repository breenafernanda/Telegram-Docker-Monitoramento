import os
import asyncio
from datetime import date, datetime, timedelta

## manipulação de excel
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from config.webdriver import iniciar_navegador
## webdrivers imports
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

## SUPORTE CHROMIUM
from selenium.webdriver.chrome.service import Service as ChromiumService
from webdriver_manager.chrome import ChromeDriverManager


## imports do monitoramento ###
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import IconSet, FormatObject, Rule
from webdriver_manager.chrome import ChromeDriverManager
import threading
import pyautogui

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import asyncio
from datetime import date, datetime, timedelta
import datetime
import json
## manipulação de excel
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from config.webdriver import iniciar_navegador
## webdrivers imports
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
## SUPORTE CHROMIUM
from selenium.webdriver.chrome.service import Service as ChromiumService
from webdriver_manager.chrome import ChromeDriverManager


## imports do monitoramento ###
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import IconSet, FormatObject, Rule
from webdriver_manager.chrome import ChromeDriverManager
import threading
import pyautogui

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PIL import Image  # Certifique-se de importar a classe Image do módulo PIL
from PIL import Image as PILImage
from telegram import InputMediaPhoto

## copiei do main
import asyncio
import logging
import os
from datetime import date, datetime
from datetime import datetime
import datetime 
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, Updater, MessageHandler, filters, ConversationHandler
from telegram import InputFile


import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ConversationHandler, MessageHandler, filters, CallbackContext
from warnings import filterwarnings
from telegram.warnings import PTBUserWarning
import asyncio
filterwarnings(action="ignore", message=r".*CallbackQueryHandler", category=PTBUserWarning)
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ConversationHandler, MessageHandler, filters, CallbackContext
import os
# %%
from openpyxl import load_workbook
import time
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import IconSet, FormatObject, Rule
from webdriver_manager.chrome import ChromeDriverManager
import threading
import pyautogui
import cv2
import numpy as np
import pyautogui
import time
import requests
import pyperclip
import pywinauto
from webdriver_manager.microsoft import EdgeChromiumDriverManager

class Driver_WPP:

    def __init__(self):
        try:
            # Define o diretório do perfil do WhatsApp Web
            dir_path = os.getcwd()
            profile = os.path.join(dir_path, "Relatórios de Geração - Whatsapp", "wpp")

            # Configura as opções do Chrome WebDriver com o perfil personalizado
            options = webdriver.ChromeOptions()

            options.add_argument(r"user-data-dir={}".format(profile))
            driver_manager = ChromeDriverManager()
            try:
                self.driver = webdriver.Chrome(executable_path=driver_manager.install(), options=options)
            except:
                self.driver = webdriver.Chrome(options=options)
        except Exception as erro: print(erro)

    async def login_whatsapp(self, update) -> None:
 
        def check_login():
            try:
                search_box = self.driver.find_element(By.XPATH, '//*[@id="pane-side"]/div[2]/div/div')
                return True
            except:
                return False
        try:
            self.driver_finish()
        except: pass
        try:

            self.driver.maximize_window()


        except Exception as e: print(e)
    
    

        print('🌐 Acessando Whatsapp Web')
        try:
            self.driver.get('https://web.whatsapp.com')
        except:
                try:
                    # Define o diretório do perfil do WhatsApp Web
                    dir_path = os.getcwd()
                    profile = os.path.join(dir_path, "Relatórios de Geração - Whatsapp", "wpp")

                    # Configura as opções do Chrome WebDriver com o perfil personalizado
                    options = webdriver.ChromeOptions()

                    options.add_argument(r"user-data-dir={}".format(profile))
                    # Crie uma instância do ChromeDriverManager e instale o driver
                    driver_manager = ChromeDriverManager()
                    self.driver = webdriver.Chrome(options=options)
                    self.driver.get('https://web.whatsapp.com')
                  
                except Exception as erro: print(erro)
        # Aguardar o login ser concluído
        
        
        time.sleep(15)
        ### tela de carregamento? Verificar !!

        # tentar obter QR CODE
        qr = True
        status = 'first'
        while qr==True:
            try:    
                # Encontre o elemento do QR code
                try:
                    self.driver.find_element(By.CSS_SELECTOR, '#app > div > div > div.landing-window > div.landing-main > div > div > div._2I5ox > div > span > button > div > span').click()

                except Exception as e:
                        pass
                try:
                    qr_code_element = self.driver.find_element(By.CSS_SELECTOR, '#app > div > div > div.landing-window > div.landing-main > div > div > div._2I5ox > div > canvas')
                except:
                    try:
                        time.sleep(10)
                        try:
                            self.driver.find_element(By.CSS_SELECTOR, '#app > div > div > div.landing-window > div.landing-main > div > div > div._2I5ox > div > span > button > div > span').click()
                        except Exception as e: 
                            qr=False 
                            # print(f'>>> qr error: {e}')
                    except:
                        qr=False

                # Obtenha as coordenadas do elemento do QR code
                left = qr_code_element.location['x']+230
                top = qr_code_element.location['y'] +35
                right = left + qr_code_element.size['width']+80
                bottom = top + qr_code_element.size['height']+120

                # Faça uma captura de tela e recorte o QR code
                self.driver.save_screenshot("screenshot.png")
                qr_code_image = Image.open("screenshot.png")
                qr_code_image = qr_code_image.crop((left, top, right, bottom))

                # Salve o QR code recortado como uma imagem PNG
                qr_code_image.save("qr_code.png")
                # print(f'>>> QR CODE GERADO E SALVO')
                with open('qr_code.png', 'rb') as arquivo:
                    qr_code_file = InputFile(arquivo)
                    # await update.message.reply_document(qr_code_file, caption='📲\nScanneie o QR Code para realizar Login...')
                    await update.message.reply_photo(qr_code_file, caption='📲\nScanneie o QR Code para realizar Login...')
                    time.sleep(25)


                    # Edite a mensagem para atualizar a foto
                ####################################
            
            except Exception as e: print(f'🚨       Erro ao obter QR Code: {e}')
        # print('Saiu do QR CODE')
        await update.message.reply_text(f'⏳ Carregando Conversas...')
        while not check_login():
            time.sleep(1)
            print('.', end="")
        print('\n ✅ Login Realizado! ✅ ')
        await update.message.reply_text("\n ✅ Login Realizado! ✅ ")


        return self.driver

    async def enviar_mensagem_wpp(self, update) -> None:
        # print('Em desenvolvimento...')
        def copiar_texto_ctrl_c(texto):

            # Copia a imagem para a área de transferência
            pyperclip.copy(texto)

            print(" 💾  Texto copiado para a área de transferência (Ctrl+C)")

        # função genérica para clicar em uma imagem 
        def clicar_img(caminho_imagem):
            try:
                # Caminho para o arquivo do print do botão
                button_image_path = caminho_imagem

                # Carrega a imagem do botão a ser localizado
                button_image = cv2.imread(button_image_path, cv2.IMREAD_UNCHANGED)

                # Verifica se a imagem foi carregada corretamente
                if button_image is None:
                    print(f"Erro ao carregar a imagem: {button_image_path}")
                    exit()

                # Captura a tela
                screenshot = pyautogui.screenshot()

                # Converte a imagem capturada para escala de cinza
                gray_screenshot = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2GRAY)

                # Realiza a correspondência de template
                result = cv2.matchTemplate(gray_screenshot, cv2.cvtColor(button_image, cv2.COLOR_BGRA2GRAY), cv2.TM_CCOEFF_NORMED)

                # Define um threshold para a correspondência
                threshold = 0.8

                # Obtém as posições onde a correspondência atende ao threshold
                locations = np.where(result >= threshold)

                # Verifica se foram encontradas correspondências
                if len(locations[0]) > 0:
                    # Obtém a posição do primeiro botão encontrado
                    if caminho_imagem == 'acessar_wpp_web.png':
                        button_position = (locations[1][0]+50, locations[0][0]+20)
                        # print(f'   ↗        {button_image_path}')

                    if caminho_imagem == 'iniciar_conversa.png':
                        button_position = (locations[1][0]+100, locations[0][0]+20)
                        # print(f'   ↗        {button_image_path}')

                    if caminho_imagem == 'quadrado.png':
                        button_position = (locations[1][0]+45, locations[0][0]+105)
                        # print(f'   ↗        {button_image_path}')

                    if caminho_imagem == 'abrir_url.png':
                        button_position = (locations[1][0]+100, locations[0][0]+35)
                        # print(f'   ↗        {button_image_path}')
                    if caminho_imagem == 'menu_anexos.png':
                        button_position = (locations[1][0]+20, locations[0][0]+5)
                        # print(f'   ↗        {button_image_path}')
                    if caminho_imagem == 'fotos2.png':
                        button_position = (locations[1][0]+5, locations[0][0]+5)
                        # print(f'   ↗        {button_image_path}')


                    # Clica no botão
                    pyautogui.click(button_position)



                return True
            
            except Exception as e:
                print(f'>>> [erro ao clicar na imagem {caminho_imagem}]: {e}')
                return False

        def verifica_existencia_imagem(caminho_imagem):
            try:
                # Caminho para o arquivo do print do botão
                button_image_path = caminho_imagem

                # Carrega a imagem do botão a ser localizado
                button_image = cv2.imread(button_image_path, cv2.IMREAD_UNCHANGED)

                # Verifica se a imagem foi carregada corretamente
                if button_image is None:
                    print(f"Erro ao carregar a imagem: {button_image_path}")
                    exit()

                # Captura a tela
                screenshot = pyautogui.screenshot()

                # Converte a imagem capturada para escala de cinza
                gray_screenshot = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2GRAY)

                # Realiza a correspondência de template
                result = cv2.matchTemplate(gray_screenshot, cv2.cvtColor(button_image, cv2.COLOR_BGRA2GRAY), cv2.TM_CCOEFF_NORMED)

                # Define um threshold para a correspondência
                threshold = 0.5

                # Obtém as posições onde a correspondência atende ao threshold
                locations = np.where(result >= threshold)




                # print(f'   🔎        {button_image_path}')
                return True
            
            except Exception as e:
                print(e)
                return False

        # verifica se a pop_up está aparecendo
        def verifica_pop_up():
            try:
                verifica = verifica_existencia_imagem('sempre_permitir.png')
                print(verifica)
                clicar_img('quadrado.png')
                clicar_img('abrir_url.png')
                try:
                    clicar_img('abrir_url.png')
                except Exception as e: print(e)

                time.sleep(2)

            except Exception as e: print(e)

        # ler arquivo Base.xslx com nomes e telefones
        def arquivo_base():
            # Carregar a planilha
            print('📃       Lendo arquivo relatorio_detalhes.xlsx')

                        # Diretório atual do script
            current_directory = os.path.dirname(os.path.abspath(__file__))
            current_directory = os.path.join(current_directory)

            # Construindo o caminho para a pasta 'relatorios'
            relatorios_directory = os.path.join(current_directory, 'relatorios')
            print('Relatorios Directory: ', relatorios_directory)
            try:
                import datetime
                # Obtenha a data e hora atual
                agora = datetime.datetime.now()

                # Formate a data e hora em 'mmYYYY'
                mes_ano = agora.strftime('%m%Y')
            except Exception as e:
                print('Erro:', e)

            # Construindo o caminho para a pasta (mesano)
            mes_ano_directory = os.path.join(relatorios_directory, mes_ano)

            # Caminho completo para o arquivo XLSX
            xlsx_file_path = os.path.join(mes_ano_directory, 'relatorio_detalhes.xlsx')
            print(xlsx_file_path)
            workbook = load_workbook(xlsx_file_path,data_only = True)
            sheet = workbook.active

            # Extrair informações de cada linha
            base_dados = []
            row = 1  # Começar na segunda linha

            while True:
                nome = sheet.cell(row=row, column=1).value
                telefone = sheet.cell(row=row, column=7).value
                email = sheet.cell(row=row, column=4).value
                mes = sheet.cell(row=row, column=5).value
                status_usina = sheet.cell(row=row, column=6).value
                instalacao = sheet.cell(row=row, column=8).value
                # mensagem_celula = sheet.cell(row=row, column=4)
                desempenho = sheet.cell(row=row, column=2).value
                # mensagem_value = mensagem_celula.value
                print(nome, telefone, email)

                if nome is None:
                    break

                # Define o nome do arquivo PDF com base na condição
                arquivo_pdf = f"{nome}.pdf"

                now = datetime.datetime.now()
                mes_ano = now.strftime('%m%Y')
                dir_path = os.getcwd()

                novo_caminho_dos_pdfs = os.path.join(dir_path,'config', "PDFs", mes_ano)


                caminho_pdf = os.path.join(novo_caminho_dos_pdfs, arquivo_pdf)
                # print('💾: ',caminho_pdf)
                    
                # Verifica se o arquivo PDF existe antes de adicioná-lo à lista
                if os.path.isfile(caminho_pdf):
                    # print('Arquivo PDF localizado! ')
                    pdf_loc = True
                else:
                    print(f"Arquivo PDF não encontrado: {caminho_pdf}")
                    pdf_loc = False
                
                def obter_mes(mes):
                    
                    try:
                        try:
                            mes = int(mes)
                        except: pass
                        if mes == 1:
                            mes_texto = 'Janeiro'
                        elif mes == 2:
                            mes_texto = 'Fevereiro'
                        elif mes == 3:
                            mes_texto = 'Março'
                        elif mes == 4:
                            mes_texto = 'Abril'
                        elif mes == 5:
                            mes_texto = 'Maio'
                        elif mes == 6:
                            mes_texto = 'Junho'
                        elif mes == 7:
                            mes_texto = 'Julho'
                        elif mes == 8:
                            mes_texto = 'Agosto'
                        elif mes == 9:
                            mes_texto = 'Setembro'
                        elif mes == 10:
                            mes_texto = 'Outubro'
                        elif mes == 11:
                            mes_texto = 'Novembro'
                        elif mes == 12:
                            mes_texto = 'Dezembro'
                        else:
                            mes_texto = 'Mês inválido'
                        return mes_texto
                    except Exception as error:
                        return f"Erro: {error}"
                
                mes_texto = obter_mes(mes)
                
                mensagem_value = (
                    
                        f"👋 Olá *{nome}*!! \n\n"
                        f"👨‍💻 Nossa equipe de especialistas realizou a análise da geração de energia do seu sistema fotovoltaico durante o mês de *{mes_texto}*! 🗓️ \n\n"
                        
                        "📃  Confira seu *Relatório de Geração Mensal* em anexo, neste chat.\n\n"
                        "🌞  Qualquer dúvida, estamos sempre à disposição! 🤝\n\n"
                        "         *KINSOL - ENERGIAS RENOVÁVEIS*     "
                    )


                # mensagem_value = (
                    
                #         f"👋 Olá *{nome}*!! \n\n"
                #         f"   *Notamos que o seu relatório do mês de agosto voltou no e-mail, então estou te encaminhando novamente por aqui* 😊\n\n"
                #         f"👨‍💻 Nossa equipe de especialistas realizou a análise da geração de energia do seu sistema fotovoltaico durante o mês de *{mes_texto}*! 🗓️ \n\n"
                        
                #         "📃  Confira seu *Relatório de Geração Mensal* em anexo, neste chat.\n\n"
                #         "🌞  Qualquer dúvida, estamos sempre à disposição! 🤝\n\n"
                #         "         *KINSOL - ENERGIAS RENOVÁVEIS*     "
                #     )
                # base_dados.append([nome, telefone, email, mensagem_value, caminho_imagem, caminho_pdf, pdf_loc])
                # Cria o caminho completo para o arquivo PDF
                base_dados.append([nome, telefone, email, mensagem_value,caminho_pdf, pdf_loc, status_usina, desempenho, instalacao,mes_texto])

                row += 1

            # Fechar a planilha
            workbook.close()
            return base_dados

        async def sender_whatsapp(update):
                    # Função para criar uma nova planilha Excel
            def criar_planilha(nome_arquivo):
                workbook = openpyxl.Workbook()

                workbook.save(nome_arquivo)
            
            def aplicar_filtro(nome_arquivo, planilha_nome='Sheet', coluna_filtro='B'):
                # Carregue a planilha existente
                workbook = load_workbook(nome_arquivo)
                
                # Acesse a planilha desejada
                planilha = workbook[planilha_nome]
                
                # Defina um filtro na coluna especificada (altere 'B' para a coluna desejada)
                planilha.auto_filter.ref = f"{coluna_filtro}:{coluna_filtro}"
                
                # Salve a planilha com o filtro aplicado
                workbook.save(nome_arquivo)

            # Função para adicionar um novo usuário verificado à planilha
            def adicionar_usuario_verificado(nome_arquivo, usuario_verificado):
                workbook = openpyxl.load_workbook(nome_arquivo)
                sheet = workbook.active

                # Adicionar os dados do usuário verificado como uma nova linha na planilha
                sheet.append(usuario_verificado)

                workbook.save(nome_arquivo)
            
            criar_planilha('Detalhes - Envio por whatsapp.xlsx')
            
            aplicar_filtro('Detalhes - Envio por whatsapp.xlsx',coluna_filtro='B')
            
            executar = True
            if executar:
                verificar_pop_up = verifica_pop_up()
                # iniciar envio de mensagens
                contatos = arquivo_base()
                # print(f'Contatos: {contatos}')
                verifica_pop_up()
                relatorios_enviados = []
                
                for contato in contatos:
                    nome_do_destinatario = contato[0]
                    telefone_do_destinatario = contato[1]
                    email = contato[2]
                    mensagem = contato[3]
                    caminho_imagem = contato[4]
                    status = contato[5]
                    status_usina = contato[6]
                    desempenho = contato[7]
                    instalacao = contato[8]
                    mes = contato[9]

                    print('\n➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖\n')
              
                    print(f'🔎{mes} {nome_do_destinatario} {email}  \n🔎 PDF ->  {status}   instalacao -> {instalacao}')
                    if status == True:

                        self.driver.get(f'https://api.whatsapp.com/send/?phone=55{telefone_do_destinatario}&text&type=phone_number&app_absent=0')
                    
                        try:
                            time.sleep(2)
                            # print('time aceito!')
                        except Exception as e: print(f'>>> [erro no time]: {e}')
                        try:
                            link_incorreto = self.driver.find_element(By.CSS_SELECTOR, '#main_block > div:nth-child(1) > h2').text
                        except Exception as e:
                            print(f'Erro ao obter o link incorreto: {e}')
                            link_incorreto = None
                        if link_incorreto =='O link está incorreto. Feche essa janela e tente usar outro link.':
                            print(link_incorreto)
                            pass
                        else:
                            print(f'Acessando conversa de {nome_do_destinatario} numero: {telefone_do_destinatario}')
                            time.sleep(1)
                            verifica_pop_up()

                            while not clicar_img('iniciar_conversa.png'):
                                try:
                                    # print('Clicar em iniciar conversa ', end=".")
                                    time.sleep(1)

                                except Exception as erro: print(f'>>>[Erro ao iniciar conversa]: {erro}')
                            web_acess = False
                            
                            while web_acess==False: 
                                try:
                                    self.driver.find_element(By.XPATH, '//*[@id="fallback_block"]/div/div/h4[2]/a/span').click()
                                    web_acess = True
                                except:
                                    # clicar_img('iniciar_conversa.png')
                                    try:
                                        self.driver.find_element(By.ID, 'action-button').click()

                                        self.driver.find_element(By.XPATH, '//*[@id="fallback_block"]/div/div/h4[2]/a/span').click()
                                        web_acess = True
                                    except:
                                        web_acess = False
                            time.sleep(5)
                        # foi pra url da conversa do cliente
                        try:
                            alerta = self.driver.find_element(By.CSS_SELECTOR, '#app > div > span:nth-child(2) > div > span > div > div > div > div > div > div.p357zi0d.ns59xd2u.kcgo1i74.gq7nj7y3.lnjlmjd6.przvwfww.mc6o24uu.e65innqk.le5p0ye3 > div > button')
                            print('Alerta de Número inválido!')
                            time.sleep(15)
                            alerta.click()
                            status_envio = '❌   Não enviado  - Número Inválido!'
                            await update.message.reply_text(f'❌ Cliente: {nome_do_destinatario}\n 🌞 {status_usina}\n ⚠️🔴 Número inválido: {telefone_do_destinatario}')
                        except:
                            if link_incorreto =='O link está incorreto. Feche essa janela e tente usar outro link.':
                                print('Alerta de Link inválido!')
                                status_envio = '❌   Não enviado  - Link Inválido!'
                                await update.message.reply_text(f'❌ Cliente: {nome_do_destinatario}\n 🌞 {status_usina}\n ⚠️🔴 Link inválido: {telefone_do_destinatario}')
                            else: 
                                # baixar a imagem e deixar no CTRL+C 
                                botao_mais = False
                                while botao_mais == False:
                                    try:
                                        try:
                                            try:
                                                self.driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div/div/div/div/span').click()
                                            except:
                                                self.driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div/div/div').click()
                                            
                                            wpp =  'wpp'
                                        except:

                                            self.driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div[2]/div/div').click()
                                            wpp = 'bussiness'
                                        button_mais = True
                                        break
                                    except Exception as e:
                                        button_mais = False

                                enviar_foto = False
                                while enviar_foto == False:
                                    try:
                                        if wpp == 'wpp':
                                            self.driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div/div/span/div/ul/div/div[1]/li/div/span').click()
                                            enviar_foto = True

                                        else:
                                            self.driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div[2]/div/span/div/div/ul/li[1]/button/span').click()
                                            enviar_foto = True
                                    except:
                                        enviar_foto = False
                                time.sleep(2)
                                selecionar_imagem = False
                                while selecionar_imagem == False:
                                    try:
                                        pyautogui.write(caminho_imagem)
                                        time.sleep(2)
                                        pyautogui.press('enter')
                                        time.sleep(1)
                                        selecionar_imagem = True
                                    except:
                                        selecionar_imagem = False
                                            # enviar_button = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[3]/div[2]/span/div/span/div/div/div[2]/div/div[2]/div[2]/div/div/span').click()

                                time.sleep(2)
                                enviar_mensagem = False

                                # copiar_texto_ctrl_c(mensagem)
                                async def enviar_msg_telegram(mensagem,caminho_imagem, update):
                                    try:
                                        # await update.message.reply_text(mensagem)
                                                                        # Verifica se o arquivo existe
                                        if os.path.exists(caminho_imagem):
                                            # Cria um objeto InputFile com o caminho do arquivo
                                        
                                            with open(caminho_imagem, 'rb') as arquivo:
                                                xlsx_input_file = InputFile(arquivo)
                                                await update.message.reply_document(xlsx_input_file, caption=mensagem)

                                            # await update.message.reply_document(xlsx_input_file, caption='✔️📊 Planilha enviada com sucesso!')
                                        else:
                                            await update.message.reply_text('❌ Arquivo não encontrado')

                                    except Exception as e: print(e)
                                
                                while enviar_mensagem  == False:
                                    try:
                                        print(f'🪪  Cliente:    {nome_do_destinatario}\n📞  Telefone:   {telefone_do_destinatario}')
                                        mensagem_telegram = (f'🪪 Cliente: {nome_do_destinatario}\n📞 Telefone: {telefone_do_destinatario}')
                                        await enviar_msg_telegram(mensagem_telegram,caminho_imagem, update)
                                        
                                        # Digitar o caminho da imagem na janela "Abrir"
                                        pyperclip.copy(mensagem)  # Copia o caminho para a área de transferência
                                        pyautogui.hotkey("ctrl", "v")  # Cole o caminho na janela "Abrir"
                                        time.sleep(1)  # Aguarde um momento para garantir que o caminho seja colado
                                        pyautogui.press("enter")  # Pressione a tecla Enter para abrir a imagem
                                        time.sleep(8)
                                        wait = True
                                        while wait == True:
                                            try:
                                                ticks_element = self.driver.find_element(By.CSS_SELECTOR, "#main > div._3B19s > div > div._5kRIK > div.n5hs2j7m.oq31bsqd.gx1rr48f.qh5tioqs > div:nth-child(9) > div > div > div > div._1BOF7._2AOIt > div.cm280p3y.m3h9lho3.lna84pfr.psacz3a6.f83pkj4x.mmw11n2j > div.lhggkp7q.ou6eaia9.qw4steeu > div > div > span")
                                                element = ticks_element.get_attribute("class")
                                                label = ticks_element.get_attribute("aria-label")
                                                print(f'label =${label}$, element = {element}')

                                                if element == 'ajgik1ph' or element == 'do8e0lj9 l7jjieqr k6y3xtnu' or label==' Entregue ':
                                                    print('✅ Mensagem enviada')
                                                    wait = False
                                                else:
                                                    print(f'🕛', end="")
                                                    wait= True # continua aguardando
                                                    print(element)
                                                    try:
                                                        elemento = self.driver.find_element(By.CSS_SELECTOR, '#main > div._3B19s > div > div._5kRIK > div.n5hs2j7m.oq31bsqd.gx1rr48f.qh5tioqs > div:nth-child(10) > div > div > div.UzMP7._3XzVs > div._1BOF7._2AOIt > div.cm280p3y.m3h9lho3.lna84pfr.psacz3a6.f83pkj4x.mmw11n2j > div.lhggkp7q.ou6eaia9.qw4steeu > div > div')
                                                        element = elemento.get_attribute("class")
                                                        label = elemento.get_attribute("aria-label")
                                                        print(f'label =${label}$, element = {element}')
                                                        if element == 'ajgik1ph' or element == 'do8e0lj9 l7jjieqr k6y3xtnu':
                                                            print('✅ Mensagem enviada')
                                                            wait = False
                                                    except Exception as e: print(f'>>> Erro no Bussiness - {e}')
                        
                                            except Exception as e: 
                                                try:
                                                    print('Não foi possível validar a mensagem (Bussiness).. Aguardando alguns segundos...')
                                                    time.sleep(8)
                                                    print('✅ Mensagem enviada')
                                                    wait = False                                     
                                                except Exception as e: 
                                                    print(f'Erro ao obter status da mensagem: {e}')

                                            time.sleep(1)

                                        enviar_mensagem = True
                                        status_envio = '✅🟢    Enviado por Whatsapp com Sucesso!'
                                        print('✅🟢    Enviado por Whatsapp com Sucesso!')
                                    except:
                                        enviar_mensagem = False
                                # print('Adicionar forma para verificar se mensagem foi enviada com sucesso!')
                                time.sleep(2)
                        
                    else:
                        try:
                            desempenho = float(desempenho)
                        except:
                            print(desempenho)
                            desempenho = -1

                        if float(desempenho)>= 72:
                            if instalacao == 'VERDADEIRO' or instalacao == True:
                                await update.message.reply_text(f'❌ Cliente: {nome_do_destinatario}\n 🌞 {status_usina}\n ⚠️🟢 PDF não localizado')
                                status_envio = '❌    Não Enviado - PDF não localizado!'
                            if instalacao == 'FALSO' or instalacao == False:
                                await update.message.reply_text(f'❌ Cliente: {nome_do_destinatario}\n 🌞 {status_usina}\n ⚠️📅 Usina instalada após data inicial!')
                                status_envio = '❌    Não Enviado - Instalação após data inicial!'
                        if desempenho == -1:
                                await update.message.reply_text(f'❌ Cliente: {nome_do_destinatario}\n 🌞 {status_usina}\n ⚠️📅 Usina instalada após data inicial!')
                                status_envio = '❌    Não Enviado - Instalação após data inicial!'

                        else:

                            await update.message.reply_text(f'❌ Cliente: {nome_do_destinatario}\n 🌞 {status_usina}\n ⚠️🔴 Geração Abaixo do esperado {desempenho} %')
                            status_envio = '❌  Não Enviado - Geração abaixo do esperado!'
                    
                    
                    verificacao = [nome_do_destinatario, status_envio, status_usina, telefone_do_destinatario, email, caminho_imagem]
                    

                    adicionar_usuario_verificado('Detalhes - Envio por whatsapp.xlsx', verificacao)
                    relatorios_enviados.append(verificacao)

                self.driver.close()
                self.driver.quit()      
        await sender_whatsapp(update)
        # Verifica se o arquivo existe
        xlsx_file_path = 'Detalhes - Envio por whatsapp.xlsx'
        if os.path.exists(xlsx_file_path):
            # Cria um objeto InputFile com o caminho do arquivo
        
            with open(xlsx_file_path, 'rb') as arquivo:
                xlsx_input_file = InputFile(arquivo)
                await update.message.reply_document(xlsx_input_file, caption='✔️📊 Detalhamento de envio por WhatsApp!')

            # await update.message.reply_document(xlsx_input_file, caption='✔️📊 Planilha enviada com sucesso!')
        
        else:
            await update.message.reply_text('❌ Arquivo não encontrado')
        
